from openpyxl import load_workbook

wb = load_workbook('/Users/giacomo/dev/deskinator-code/tests/results/no-offset/robot_test_results_no_offset.xlsx')
ws = wb['Summary']

print('Summary Statistics:')
print('-' * 80)
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if row[0]:
        if i == 0:  # Header row
            print(f'{row[0]:<25} {row[1]:>12} {row[2]:>12} {row[3]:>12} {row[4]:>12}')
        else:  # Data rows
            print(f'{row[0]:<25} {row[1]:>12.2f} {row[2]:>12.2f} {row[3]:>12.2f} {row[4]:>12.2f}')
