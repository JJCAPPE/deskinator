"""
Visualize robot testing results with Gaussian distribution fits.

Reads trial data from Excel file and generates distribution plots
with fitted Gaussian curves for key performance metrics.

Usage:
    python visualize_results.py tests/results/results.xlsx
"""

import numpy as np
import matplotlib.pyplot as plt
from scipy import stats
from openpyxl import load_workbook
import argparse
from pathlib import Path


def read_excel_data(excel_path: str) -> dict:
    """
    Read trial data from Excel file.
    
    Args:
        excel_path: Path to Excel file with trial data
    
    Returns:
        Dictionary with metric names as keys and lists of values
    """
    wb = load_workbook(excel_path)
    ws = wb['Trial Data']
    
    # Read headers (skip trial number column)
    headers = []
    for cell in ws[1][1:]:  # Skip first column (Trial)
        headers.append(cell.value)
    
    # Read data rows
    data = {header: [] for header in headers}
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        trial_num = row[0]
        for i, value in enumerate(row[1:]):
            if value is not None:
                data[headers[i]].append(value)
    
    return data


def plot_gaussian_distribution(values, metric_name, output_path, bins=20):
    """
    Plot histogram with fitted Gaussian distribution.
    
    Args:
        values: List of metric values
        metric_name: Name of the metric for title
        output_path: Path to save the figure
        bins: Number of histogram bins
    """
    values = np.array(values)
    
    # Calculate statistics
    mean = np.mean(values)
    std = np.std(values)
    
    # Create figure
    fig, ax = plt.subplots(figsize=(10, 6))
    
    # Plot histogram
    n, bins_arr, patches = ax.hist(values, bins=bins, density=True, 
                                     alpha=0.7, color='steelblue', 
                                     edgecolor='black', label='Data')
    
    # Fit and plot Gaussian
    x = np.linspace(values.min(), values.max(), 200)
    gaussian = stats.norm.pdf(x, mean, std)
    ax.plot(x, gaussian, 'r-', linewidth=2.5, label=f'Gaussian fit\nμ={mean:.2f}, σ={std:.2f}')
    
    # Add mean line
    ax.axvline(mean, color='darkred', linestyle='--', linewidth=2, 
               label=f'Mean: {mean:.2f}')
    
    # Add ±1σ lines
    ax.axvline(mean - std, color='orange', linestyle=':', linewidth=1.5, 
               label=f'μ ± σ: [{mean-std:.2f}, {mean+std:.2f}]')
    ax.axvline(mean + std, color='orange', linestyle=':', linewidth=1.5)
    
    # Labels and title
    ax.set_xlabel(metric_name, fontsize=12, fontweight='bold')
    ax.set_ylabel('Probability Density', fontsize=12, fontweight='bold')
    ax.set_title(f'Distribution of {metric_name}\n({len(values)} trials)', 
                 fontsize=14, fontweight='bold', pad=20)
    
    # Legend
    ax.legend(loc='upper right', fontsize=10, framealpha=0.9)
    
    # Grid
    ax.grid(True, alpha=0.3, linestyle='--')
    
    # Tight layout
    plt.tight_layout()
    
    # Save figure
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"  ✓ Saved: {output_path}")


def create_time_comparison_plot(boundary_times, coverage_times, total_times, output_path):
    """
    Create a comparison plot for all three time metrics.
    
    Args:
        boundary_times: List of boundary discovery times
        coverage_times: List of coverage times
        total_times: List of total times
        output_path: Path to save the figure
    """
    fig, axes = plt.subplots(1, 3, figsize=(18, 5))
    
    time_data = [
        (boundary_times, 'Boundary Time (s)', axes[0]),
        (coverage_times, 'Coverage Time (s)', axes[1]),
        (total_times, 'Total Time (s)', axes[2])
    ]
    
    for values, title, ax in time_data:
        values = np.array(values)
        mean = np.mean(values)
        std = np.std(values)
        
        # Histogram
        n, bins_arr, patches = ax.hist(values, bins=15, density=True,
                                        alpha=0.7, color='steelblue',
                                        edgecolor='black')
        
        # Gaussian fit
        x = np.linspace(values.min(), values.max(), 200)
        gaussian = stats.norm.pdf(x, mean, std)
        ax.plot(x, gaussian, 'r-', linewidth=2.5)
        
        # Mean line
        ax.axvline(mean, color='darkred', linestyle='--', linewidth=2)
        
        # Labels
        ax.set_xlabel(title, fontsize=11, fontweight='bold')
        ax.set_ylabel('Probability Density', fontsize=11, fontweight='bold')
        ax.set_title(f'{title}\nμ={mean:.2f}s, σ={std:.2f}s', 
                     fontsize=12, fontweight='bold')
        ax.grid(True, alpha=0.3, linestyle='--')
    
    plt.suptitle('Time Metrics Distribution Comparison', 
                 fontsize=16, fontweight='bold', y=1.02)
    plt.tight_layout()
    
    # Save figure
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    plt.close()
    
    print(f"  ✓ Saved: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Visualize robot testing results with Gaussian fits"
    )
    
    parser.add_argument(
        "excel_file",
        type=str,
        help="Path to Excel file with trial results"
    )
    
    parser.add_argument(
        "--output-dir",
        type=str,
        default=None,
        help="Output directory for graphs (default: same directory as Excel file + /graphs)"
    )
    
    args = parser.parse_args()
    
    # Determine output directory
    excel_path = Path(args.excel_file)
    if args.output_dir:
        output_dir = Path(args.output_dir)
    else:
        output_dir = excel_path.parent / "graphs"
    
    # Create output directory
    output_dir.mkdir(parents=True, exist_ok=True)
    
    print("=" * 60)
    print("Robot Testing Results Visualization")
    print("=" * 60)
    print(f"Input file: {excel_path}")
    print(f"Output directory: {output_dir}")
    print("=" * 60)
    
    # Read data
    print("\nReading data from Excel...")
    data = read_excel_data(str(excel_path))
    print(f"  ✓ Loaded {len(data['Edge Points'])} trials")
    
    # Generate individual plots
    print("\nGenerating distribution plots...")
    
    # 1. Rectangle fitting error
    plot_gaussian_distribution(
        data['Rect Error (ppm)'],
        'Rectangle Fitting Error (%)',
        output_dir / 'rect_error_distribution.png',
        bins=20
    )
    
    # 2. Coverage (full rectangle)
    plot_gaussian_distribution(
        data['Coverage Full (%)'],
        'Coverage Full (%)',
        output_dir / 'coverage_full_distribution.png',
        bins=20
    )
    
    # 3. Coverage (inset rectangle)
    plot_gaussian_distribution(
        data['Coverage Inset (%)'],
        'Coverage Inset (%)',
        output_dir / 'coverage_inset_distribution.png',
        bins=20
    )
    
    # 4. Distance covered
    plot_gaussian_distribution(
        data['Distance (m)'],
        'Distance Covered (m)',
        output_dir / 'distance_distribution.png',
        bins=20
    )
    
    # 5. Time metrics comparison
    print("\nGenerating time comparison plot...")
    create_time_comparison_plot(
        data['Boundary Time (s)'],
        data['Coverage Time (s)'],
        data['Total Time (s)'],
        output_dir / 'time_comparison.png'
    )
    
    # Individual time plots
    plot_gaussian_distribution(
        data['Boundary Time (s)'],
        'Boundary Discovery Time (s)',
        output_dir / 'boundary_time_distribution.png',
        bins=15
    )
    
    plot_gaussian_distribution(
        data['Coverage Time (s)'],
        'Coverage Time (s)',
        output_dir / 'coverage_time_distribution.png',
        bins=15
    )
    
    plot_gaussian_distribution(
        data['Total Time (s)'],
        'Total Time (s)',
        output_dir / 'total_time_distribution.png',
        bins=15
    )
    
    print("\n" + "=" * 60)
    print("Visualization complete!")
    print("=" * 60)
    print(f"\nGenerated {len(list(output_dir.glob('*.png')))} plots in:")
    print(f"  {output_dir}")
    print("\nPlots created:")
    print("  - rect_error_distribution.png")
    print("  - coverage_full_distribution.png")
    print("  - coverage_inset_distribution.png")
    print("  - distance_distribution.png")
    print("  - time_comparison.png")
    print("  - boundary_time_distribution.png")
    print("  - coverage_time_distribution.png")
    print("  - total_time_distribution.png")
    print("=" * 60)


if __name__ == "__main__":
    main()
