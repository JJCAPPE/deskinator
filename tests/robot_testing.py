"""
Robot performance testing framework.

Runs multiple trials of table cleaning simulation and collects
performance metrics. Exports results to Excel for analysis.

Usage:
    python robot_testing.py --trials 25 --output results.xlsx
"""

import numpy as np
import argparse
import time
from typing import List, Tuple, Optional, Dict
import sys
from pathlib import Path

# Add parent directory to path for imports when running from tests/
parent_dir = Path(__file__).parent.parent
if str(parent_dir) not in sys.path:
    sys.path.insert(0, str(parent_dir))
# Add tests directory to path for local imports
tests_dir = Path(__file__).parent
if str(tests_dir) not in sys.path:
    sys.path.insert(0, str(tests_dir))

# Import our modules
from planning.coverage import (
    SimpleWallFollower,
    SimpleRectangleFit,
    CoveragePlanner,
)
from planning.map2d import SweptMap
from utils.viz import Visualizer
from config import GEOM, ALG, LIMS
from analysis_utils import calculate_rectangle_error, calculate_trajectory_distance

# Import viz_demo components for robot simulation
from viz_demo import (
    SimulatedRobot,
    TableSimulator,
    simulate_wall_following,
    simulate_coverage,
)


# Sentinel object to indicate visualization should be disabled
_NO_VIZ = object()


def run_single_trial(
    trial_num: int,
    speed_multiplier: float = 2.0,
    verbose: bool = False,
    viz: Optional[Visualizer] = _NO_VIZ,
) -> Dict:
    """
    Run a single trial and collect all metrics.

    Args:
        trial_num: Trial number for logging
        speed_multiplier: Speed multiplier for simulation (higher = faster, default: 2.0)
        verbose: Print progress messages
        viz: Optional Visualizer instance for real-time visualization

    Returns:
        Dictionary with all collected metrics
    """
    if verbose:
        print(f"\nTrial {trial_num}:")
        print("-" * 40)

    # Create or reuse visualizer (passed from main)
    # If viz is None, we pass None to simulation functions (no visualization)

    # Create swept map
    swept_map = SweptMap()

    # === PHASE 1: Boundary Discovery ===
    t_boundary_start = time.time()

    # Run wall following simulation
    robot, wall_follower, rect_fit, table, _, swept_map, boundary_sim_time = (
        simulate_wall_following(
            speed_multiplier=speed_multiplier,
            viz=viz,
            swept_map=swept_map,
            verbose=verbose,
        )
    )

    boundary_time = time.time() - t_boundary_start
    rectangle = rect_fit.get_rectangle()

    if verbose:
        print(
            f"  Boundary discovery: {boundary_sim_time:.1f}s (wall time: {boundary_time:.2f}s)"
        )
        print(f"  Edge points: {len(rect_fit.edge_points)}")

    # Track trajectory at boundary completion
    boundary_trajectory_length = len(robot.trajectory)

    # === PHASE 2: Coverage ===
    coverage_success = False
    coverage_time_real = 0.0
    coverage_sim_time = 0.0
    planner = None

    if rectangle:
        t_coverage_start = time.time()

        # Run coverage simulation
        coverage_success, coverage_sim_time, planner = simulate_coverage(
            robot=robot,
            rect_fit=rect_fit,
            table=table,
            speed_multiplier=speed_multiplier,
            viz=viz,
            swept_map=swept_map,
            verbose=verbose,
        )

        coverage_time_real = time.time() - t_coverage_start

        if verbose:
            print(
                f"  Coverage: {coverage_sim_time:.1f}s (wall time: {coverage_time_real:.2f}s)"
            )
            print(f"  Complete: {coverage_success}")

    # === Collect Metrics ===
    metrics = {}

    # 1. Points on edge found
    metrics["edge_points"] = len(rect_fit.edge_points)

    # 2. Rectangle fitting error (%)
    ground_truth = table.get_ground_truth_rect()
    metrics["rect_error_pct"] = calculate_rectangle_error(rectangle, ground_truth)

    # 3. Coverage % (full rectangle)
    if rectangle:
        metrics["coverage_full_pct"] = swept_map.coverage_ratio(rectangle) * 100.0
    else:
        metrics["coverage_full_pct"] = 0.0

    # 4. Coverage % (inset rectangle)
    if rectangle:
        planner_temp = CoveragePlanner()
        planner_temp.set_rectangle(rectangle)
        inset_rect = planner_temp.get_inset_rectangle()
        if inset_rect:
            metrics["coverage_inset_pct"] = swept_map.coverage_ratio(inset_rect) * 100.0
        else:
            metrics["coverage_inset_pct"] = 0.0
    else:
        metrics["coverage_inset_pct"] = 0.0

    # 5. Boundary discovery time
    metrics["boundary_time_s"] = boundary_sim_time

    # 6. Coverage time
    metrics["coverage_time_s"] = coverage_sim_time

    # 7. Total time
    metrics["total_time_s"] = boundary_sim_time + coverage_sim_time

    # 8. Distance covered
    metrics["distance_m"] = calculate_trajectory_distance(robot.trajectory)

    if verbose:
        print(f"  Metrics collected:")
        print(f"    Edge points: {metrics['edge_points']}")
        # Display rect error: show ppm, and also percentage for reference
        rect_error_ppm = metrics["rect_error_pct"]
        rect_error_pct = rect_error_ppm / 10000.0  # Convert ppm to percentage
        print(f"    Rect error: {rect_error_ppm:.2f} ppm ({rect_error_pct:.4f}%)")
        print(f"    Coverage (full): {metrics['coverage_full_pct']:.2f}%")
        print(f"    Coverage (inset): {metrics['coverage_inset_pct']:.2f}%")
        print(f"    Distance: {metrics['distance_m']:.2f}m")

    return metrics


def export_to_excel(results: List[Dict], output_path: str):
    """
    Export trial results to Excel with raw data and summary sheets.

    Args:
        results: List of metric dictionaries from trials
        output_path: Path to output Excel file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("\nError: openpyxl not installed. Install with:")
        print("  pip install openpyxl")
        return

    wb = Workbook()

    # === Sheet 1: Trial Data ===
    ws_data = wb.active
    ws_data.title = "Trial Data"

    # Headers
    headers = [
        "Trial",
        "Edge Points",
        "Rect Error (ppm)",
        "Coverage Full (%)",
        "Coverage Inset (%)",
        "Boundary Time (s)",
        "Coverage Time (s)",
        "Total Time (s)",
        "Distance (m)",
    ]

    ws_data.append(headers)

    # Format headers
    for cell in ws_data[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(
            start_color="B4C7E7", end_color="B4C7E7", fill_type="solid"
        )

    # Data rows
    metric_keys = [
        "edge_points",
        "rect_error_pct",
        "coverage_full_pct",
        "coverage_inset_pct",
        "boundary_time_s",
        "coverage_time_s",
        "total_time_s",
        "distance_m",
    ]

    for i, result in enumerate(results, start=1):
        row = [i] + [result.get(key, 0.0) for key in metric_keys]
        ws_data.append(row)

    # Set column widths
    ws_data.column_dimensions["A"].width = 8
    for col in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws_data.column_dimensions[col].width = 16

    # === Sheet 2: Summary Statistics ===
    ws_summary = wb.create_sheet("Summary")

    # Headers
    ws_summary.append(["Metric", "Mean", "Std Dev", "Min", "Max"])
    for cell in ws_summary[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(
            start_color="B4C7E7", end_color="B4C7E7", fill_type="solid"
        )

    # Calculate statistics for each metric
    metric_names = [
        "Edge Points",
        "Rect Error (ppm)",
        "Coverage Full (%)",
        "Coverage Inset (%)",
        "Boundary Time (s)",
        "Coverage Time (s)",
        "Total Time (s)",
        "Distance (m)",
    ]

    for metric_name, metric_key in zip(metric_names, metric_keys):
        values = [r[metric_key] for r in results if metric_key in r]
        if values:
            mean_val = np.mean(values)
            std_val = np.std(values)
            min_val = np.min(values)
            max_val = np.max(values)
            ws_summary.append([metric_name, mean_val, std_val, min_val, max_val])

    # Set column widths
    ws_summary.column_dimensions["A"].width = 20
    for col in ["B", "C", "D", "E"]:
        ws_summary.column_dimensions[col].width = 14

    # Save workbook
    wb.save(output_path)
    print(f"\n✓ Results exported to: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Run robot performance testing trials",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    parser.add_argument(
        "--trials",
        type=int,
        default=25,
        help="Number of trials to run (default: 25)",
    )

    parser.add_argument(
        "--output",
        type=str,
        default="tests/results/robot_test_results.xlsx",
        help="Output Excel file path (default: tests/results/robot_test_results.xlsx)",
    )

    parser.add_argument(
        "--speed",
        type=float,
        default=2.0,
        help="Speed multiplier for simulation (default: 2.0) - multiplies ALL speeds",
    )

    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print detailed progress for each trial",
    )

    parser.add_argument(
        "--no-viz",
        action="store_true",
        default=False,
        help="Disable real-time visualization (disabled by default)",
    )

    parser.add_argument(
        "--seed",
        type=int,
        default=None,
        help="Random seed for reproducibility (optional)",
    )

    args = parser.parse_args()

    # Set random seed if provided
    if args.seed is not None:
        np.random.seed(args.seed)
        print(f"Random seed set to: {args.seed}")

    print("=" * 60)
    print("Robot Performance Testing Framework")
    print("=" * 60)
    print(f"Trials: {args.trials}")
    print(f"Output: {args.output}")
    print(f"Speed multiplier: {args.speed}x (multiplies ALL speeds)")
    print(f"Visualization: {'Disabled' if args.no_viz else 'Enabled (real-time)'}")
    print("=" * 60)

    # Create shared visualizer for all trials (enabled by default, matching viz_demo.py)
    # This allows us to see each trial's progress sequentially in the same window
    viz = None
    if not args.no_viz:
        viz = Visualizer(figsize=(14, 8))

    # Run trials
    results = []
    start_time = time.time()

    for trial_num in range(1, args.trials + 1):
        if not args.verbose:
            print(f"Running trial {trial_num}/{args.trials}...", end="", flush=True)

        metrics = run_single_trial(trial_num, args.speed, args.verbose, viz=viz)
        results.append(metrics)

        if not args.verbose:
            print(" ✓")

    # Close shared visualizer if it was created
    if viz is not None:
        viz.close()

    total_time = time.time() - start_time

    print("\n" + "=" * 60)
    print(f"All {args.trials} trials completed in {total_time:.1f}s")
    print("=" * 60)

    # Export to Excel
    export_to_excel(results, args.output)

    # Print summary
    print("\n" + "=" * 60)
    print("Summary Statistics")
    print("=" * 60)

    metric_keys = [
        ("Edge Points", "edge_points"),
        ("Rect Error (ppm)", "rect_error_pct"),
        ("Coverage Full (%)", "coverage_full_pct"),
        ("Coverage Inset (%)", "coverage_inset_pct"),
        ("Boundary Time (s)", "boundary_time_s"),
        ("Coverage Time (s)", "coverage_time_s"),
        ("Total Time (s)", "total_time_s"),
        ("Distance (m)", "distance_m"),
    ]

    for name, key in metric_keys:
        values = [r[key] for r in results]
        mean_val = np.mean(values)
        std_val = np.std(values)
        # Use more decimal places for rect error (ppm)
        if key == "rect_error_pct":
            print(f"{name:25s}: {mean_val:10.2f} ± {std_val:8.2f}")
        else:
            print(f"{name:25s}: {mean_val:8.2f} ± {std_val:6.2f}")

    print("=" * 60)


if __name__ == "__main__":
    main()
